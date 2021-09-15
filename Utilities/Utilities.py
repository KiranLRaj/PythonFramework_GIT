import os
import time

from selenium.common.exceptions import TimeoutException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
import openpyxl


class Utilities:

    def __init__(self,driver):
        self.driver=driver

    def CheckElementExists(self, ElemXPath):
        WaitElement = WebDriverWait(self.driver, 3)
        try:
            WaitElement.until(expected_conditions.visibility_of_element_located((By.XPATH, ElemXPath)))
            print("Element Exist")
            return True
        except TimeoutException:
            print("Element Does Not exist")
            return False

    def ExcelDataOneColumn(self,path,sheetname):
        DataFetched = []
        DataFetched.clear()
        EcxelObj = openpyxl.load_workbook(path)
        EcxelObj._active_sheet_index = EcxelObj.sheetnames.index(sheetname)
        Sheet = EcxelObj.active
        RowNum = Sheet.max_row
        for i in range(1, RowNum + 1):
            DataFetched.append(Sheet.cell(i, 1).value)
        return DataFetched

    def TypeIntoElement(self,Element,Value):
        Actions = ActionChains(self.driver)
        Actions.click(Element).perform()
        Actions.key_down(Keys.CONTROL).send_keys("a").key_up(Keys.CONTROL).send_keys(Keys.BACK_SPACE).perform()
        time.sleep(1)
        Actions.send_keys_to_element(Element,(str(Value).strip())).perform()
        Actions.reset_actions()

    def TypeAndEnter(self,Element,Value):
        Actions = ActionChains(self.driver)
        Actions.click(Element).perform()
        Actions.key_down(Keys.CONTROL).send_keys("a").key_up(Keys.CONTROL).send_keys(Keys.BACK_SPACE).perform()
        time.sleep(1)
        Actions.send_keys_to_element(Element,(str(Value).strip() + Keys.ENTER).perform()).perform()
        Actions.reset_actions()

    def GetRGB_Name(self,Element):
        RGB=[]
        RGB_Dec=[]
        BG=Element.value_of_css_property("background-color")
        Values = (str(BG.split("(")[1]).replace(")", "")).split(",")
        self.driver.execute_script("window.open('');")
        self.driver.switch_to.window(self.driver.window_handles[1])
        self.driver.get("http://www.color-blindness.com/color-name-hue/")
        self.driver.switch_to.frame(self.driver.find_element_by_xpath("//iframe[@loading='lazy']"))
        self.TypeIntoElement(self.driver.find_element_by_xpath("//*[@id='main_frame']//*[@id='cp1_Red']"), Values[0])
        self.TypeIntoElement(self.driver.find_element_by_xpath("//*[@id='main_frame']//*[@id='cp1_Green']"), (Values[1]))
        self.TypeIntoElement(self.driver.find_element_by_xpath("//*[@id='main_frame']//*[@id='cp1_Blue']"), (Values[2]))
        time.sleep(1)
        RGB.append(self.driver.find_element_by_xpath("//*[@id='cp1_ColorNameText']").text)
        self.driver.close()
        self.driver.switch_to.window(self.driver.window_handles[0])
        for i in range(0,3):
            RGB_Dec.append(Values[i].strip())
        RGB.append(RGB_Dec)
        return RGB

    def ExcelDataMultipleColumns(self,path,sheetname):
        DataFetched = []
        DataRowWise = []
        DataFetched.clear()
        DataRowWise.clear()
        EcxelObj = openpyxl.load_workbook(path)
        EcxelObj._active_sheet_index = EcxelObj.sheetnames.index(sheetname)
        Sheet = EcxelObj.active
        RowNum = Sheet.max_row
        Column = Sheet.max_column
        for i in range(1, RowNum + 1):
            for j in range(1, Column + 1):
                DataFetched.append(Sheet.cell(i, j).value)
            DataRowWise.append(DataFetched)
            DataFetched = []
        return DataRowWise

    def WriteExcel(self,DataList,ColumnNum):
        Wb = openpyxl.Workbook()
        Ws = Wb.active
        for i in range(0, len(DataList)):
            Ws.cell(i + 1, ColumnNum).value = DataList[i]
            if i == 3:
                break
        Wb.save("C:\\Users\Raj\Desktop\\DataWrite.xlsx")





